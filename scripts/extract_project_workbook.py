#!/usr/bin/env python3
"""
Извлекает книгу «Калькулятор стоимости проектных работ ИНСИ» (раздел КМ)
в generated-TS для движка вкладки «Проектные работы».

В проекте используется ТОЛЬКО конструктив «Великан»: движок принудительно
выставляет тип здания = Великан (AM7=3) и флаг расчёта КМ (AM8=TRUE).
Прочие ангары (Спринт/Атлант/Крон) в подбор не идут — их колонки
обнуляются гейтом CHOOSE и не влияют на результат.

Источник: EXCEL-011 (см. knowledge/raw/source-register.md). Файл не
коммитится; здесь только обезличенные формулы книги.

Воспроизведение книги доказано офлайн-оракулом на HyperFormula: при
дефолтном сценарии Z9=179.2 тыс.руб, Z10=19.91 раб.дн — совпадает с
кэшированными значениями Excel.

Преобразования для HyperFormula:
  - литералы TRUE/FALSE → TRUE()/FALSE();
  - префикс _xlfn. удаляется (CEILING.MATH и т.п.).
"""
import openpyxl, json, re, sys

SRC = sys.argv[1] if len(sys.argv) > 1 else "/tmp/proj.xlsx"
OUT = "src/calc/projectWork/workbook.generated.ts"

def fix_formula(v):
    if isinstance(v, str) and v.startswith("="):
        v = re.sub(r'\bTRUE\b(?!\()', 'TRUE()', v)
        v = re.sub(r'\bFALSE\b(?!\()', 'FALSE()', v)
        v = v.replace("_xlfn.", "")
    return v

wb = openpyxl.load_workbook(SRC, data_only=False)
sheets = {}
for ws in wb.worksheets:
    grid = []
    for r in range(1, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if isinstance(v, bool):
                row.append(v)
            elif isinstance(v, (int, float)) or v is None:
                row.append(v)
            else:
                row.append(fix_formula(str(v)))
        grid.append(row)
    sheets[ws.title] = grid

# Входные ячейки (только то, чем управляет вкладка «Проектные работы»).
input_cells = {
    # геометрия (видимая зона) — общие с основным зданием
    "spanCount": "D33",          # количество пролётов поперёк
    "spanWidth1": "B35", "spanWidth2": "D35", "spanWidth3": "F35",
    "spanWidth4": "H35", "spanWidth5": "J35",
    "length": "B38",             # длина здания, м
    "framePitch": "B40",         # шаг рам, м
    "height": "F43",             # высота, м
    "seismic": "AM6",            # 1=нет, 2/3/4 = категория (7/8/9 баллов)
    "buildingType": "AM7",       # ВСЕГДА 3 = Великан
    "kmFlag": "AM8",             # ВСЕГДА TRUE — расчёт раздела КМ
    "roof": "AM11",              # 1=односк, 2=двуск, 3=плоская, 4=многоск
    "overheadCrane": "AM18", "overheadCraneCapacity": "AM19",
    "suspendedCrane": "AM20", "suspendedCraneCapacity": "AM21",
    # ограждающие конструкции
    "wallMaterial": "AM36",      # 1/2/3
    "wallThickness": "AM37",
    "roofingMaterial": "AM31",   # 1/3/4=профлист, 2=послойная сборка
    "snowGuard": "AM33",         # снегозадержание
    "roofFence": "AM34",         # ограждение кровли
    "drainage": "AM70",          # водосточная система
    "windows": "AM38", "windowCount": "W51",
    "gates": "AM40", "gateCount": "W56", "gateSpanCount": "W58",
    "doors": "AM41", "doorCount": "W61",
    # перегородки
    "partitionGvl": "AM42",
    "partitionSandwichLayered": "AM43",
    "partitionSandwichFactory": "AM44",
    "partitionGvlArea": "B81",                 # площадь перегородок ГВЛ, м²
    "partitionSandwichLayeredArea": "B85",     # площадь перегородок СП послойной, м²
    "partitionSandwichFactoryArea": "B88",     # площадь перегородок СП заводской, м²
    "partitionOpeningsDoors": "AM45",
    "partitionOpeningsGates": "AM46",
    # парапет (3=нет, 1=одна сторона, 2=обе)
    "parapetLongSide": "AM59", "parapetEnd": "AM60", "parapetOverhang": "AM61",
    # перекрытие и антресоли
    "floor": "AM14",
    "mezzanine1": "AM15", "mezzanine2": "AM16", "mezzanine3": "AM17",
    # лестницы (мастер-флаг + 8 типов + счётчики маршей)
    "stairs": "AM22",
    "stairRcSingle": "AM23", "stairRcDouble": "AM24",
    "stairRcTriple": "AM25", "stairRcQuad": "AM26",
    "stairMetalSingle": "AM27", "stairMetalDouble": "AM28",
    "stairMetalTriple": "AM29", "stairMetalQuad": "AM30",
    "stairRcSingleCount": "F71", "stairRcDoubleCount": "F73",
    "stairRcTripleCount": "F75", "stairRcQuadCount": "F77",
    "stairMetalSingleCount": "N71", "stairMetalDoubleCount": "N73",
    "stairMetalTripleCount": "N75", "stairMetalQuadCount": "N77",
    # прочее
    "fireResistance": "AM63",    # степень огнестойкости
    "overheadCost": "X2",        # издержки, доля
}
output_cells = {
    "kmCostThousandRub": "Z9",   # стоимость проектных работ КМ, тыс.руб
    "kmDurationDays": "Z10",     # срок проектирования, раб.дней
    "costPerTon": "Z17",         # стоимость, тыс.руб/т (Великан)
}

# Дефолтный сценарий = как в исходном файле, но тип принудительно Великан.
default_inputs = {"buildingType": 3, "kmFlag": True, "roof": 2}
default_expected = {"kmCostThousandRub": 195.6, "kmDurationDays": 21.733333333333334}

payload = {
    "sourceId": "EXCEL-011",
    "summarySheet": "Сроки и стоимость",
    "sheets": sheets,
    "inputCells": input_cells,
    "outputCells": output_cells,
    "defaultScenario": {"inputs": default_inputs, "expected": default_expected},
}

header = (
    "/* eslint-disable */\n"
    "// Generated by scripts/extract_project_workbook.py — НЕ редактировать вручную.\n"
    "// Источник: EXCEL-011 (калькулятор стоимости проектных работ ИНСИ, раздел КМ).\n"
    "// Используется только конструктив «Великан» (buildingType=3).\n"
)
with open(OUT, "w", encoding="utf-8") as f:
    f.write(header)
    f.write("export const projectWorkbook = ")
    f.write(json.dumps(payload, ensure_ascii=False))
    f.write(" as const;\n")

print("written", OUT)
print("sheets:", {k: f"{len(v)}x{len(v[0]) if v else 0}" for k, v in sheets.items()})
